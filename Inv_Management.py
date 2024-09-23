from tkinter import *
from tkinter import ttk, messagebox,filedialog
import subprocess
import copy
import sys
import os
import pickle
from datetime import datetime,timedelta
from collections import defaultdict
try:
    import matplotlib
except ImportError:
    print("> 'matplotlib' module is missing!\n" +"Trying to install required module: matplotlib")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "matplotlib"])
finally:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure

try:
    import PIL
except ImportError:
    print("> 'numpy' module is missing!\n" +
          "Trying to install required module: Pillow")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pillow"])
    print()
finally:
    import PIL
    from PIL import ImageTk, Image, ImageDraw

try:
    import pandas
except ImportError:
    print("> 'pandas' module is missing!\n" +
          "Trying to install required module: pandas")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
finally:
    import pandas as pd
try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("> 'openpyxl' module is missing!\n" +
          "Trying to install required module: openpyxl")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
finally:
    import openpyxl
class BasePage(Frame):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.master = master
        self.current_frame = None
        self.history_list = []
        self.history_tree = None

    def show_notification(self, message):
        messagebox.showinfo("Notification", message)

    def switch_page(self, page_class):
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = page_class(self.master)
        self.current_frame.place(x=200, y=100, width=1150, height=700)
    
    def save_data(self):
        try:
            with open("product_list.pkl", "wb") as product_file:
                pickle.dump(self.product_list, product_file)
        except FileNotFoundError:
            pass
        try:
            with open("order_list.pkl", "wb") as order_file:
                pickle.dump(self.order_list, order_file)
        except FileNotFoundError:
            pass
        try:
            with open("Sale_list.pkl", "wb") as Sale_file:
                pickle.dump(self.order_list, Sale_file)
        except FileNotFoundError:
            pass
        try:
            with open("History_list.pkl", "wb") as History_file:
                pickle.dump(self.order_list, History_file)
        except FileNotFoundError:
            pass

    def load_data(self):
        try:
            with open("product_list.pkl", "rb") as product_file:
                self.product_list = pickle.load(product_file)
        except FileNotFoundError:
            pass

        try:
            with open("order_list.pkl", "rb") as order_file:
                self.order_list = pickle.load(order_file)
        except (FileNotFoundError, EOFError):
            pass

        try:
            with open("Sale_list.pkl", "rb") as Sale_file:
                self.Sale_list = pickle.load(Sale_file)
        except FileNotFoundError:
            pass

        try:
            with open("History_list.pkl", "rb") as History_file:
                self.history_list = pickle.load(History_file)
        except FileNotFoundError:
            pass

    def add_to_history(self, action, item_name=None):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if item_name:
            entry = f"{timestamp} - {action}: {item_name}"
        else:
            entry = f"{timestamp} - {action}"

        history_file_path = "History_list.pkl"
        if not os.path.exists(history_file_path):
            self.history_list = []
        else:
            with open(history_file_path, "rb") as history_file:
                self.history_list = pickle.load(history_file)

        self.history_list.append(entry)

        with open(history_file_path, "wb") as history_file:
            pickle.dump(self.history_list, history_file)

        if self.history_tree:
            self.update_history_widget()

    def update_history_widget(self):
        if self.history_tree:
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)

            for entry in self.history_list:
                self.history_tree.insert("", "end", values=(entry,))

class HomePage(BasePage):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.history_list = []
        self.show_home_page()
        
    def show_home_page(self):
        self.load_data()
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = Frame(self.master,background="White")
        self.current_frame.place(x=200, y=100, width=1150, height=600)
        home = Label(self.current_frame, text="Home Page", font=("Arial", 24),background="White")
        home.pack(pady=20)
        try:
            with open("product_list.pkl", "rb") as product_file:
                product_data = pickle.load(product_file)

                type_quantity = defaultdict(float)

                for product in product_data:
                    type_quantity[product["Type"]] += product["Quantity"]

                merged_labels = list(type_quantity.keys())
                merged_sizes = list(type_quantity.values())
                total_quantity = sum(merged_sizes)

                colors = ['red', 'yellow', 'purple', 'orange', 'green', 'blue', 'cyan', 'magenta', 'lime', 'pink']
                fig = Figure(figsize=(6, 2),tight_layout={'pad': 1.0})
                ax = fig.add_subplot(111, aspect='equal')

                wedges, texts = ax.pie(merged_sizes, labels=None, startangle=75, colors=colors, pctdistance=0)

                center_circle = plt.Circle((0, 0), 0.70, fc='white')
                ax.add_artist(center_circle)

                ax.set_title("Product Type")

                legend_labels = [f'{label} - {size:.2f}' for label, size in zip(merged_labels, merged_sizes)]
                ax.legend(wedges, legend_labels, title="Legend", bbox_to_anchor=(1, 0.5), loc="center left", borderaxespad=0.0)

                canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
                canvas_widget.get_tk_widget().place(x=-100, y=50, width=500, height=250)
        except FileNotFoundError:
            pass

        try:
            with open("order_list.pkl", "rb") as order_file:
                order_data = pickle.load(order_file)

                type_quantity = defaultdict(float)

                for order in order_data:
                    type_quantity[order["Name"]] += order["Quantity"]

                merged_labels = list(type_quantity.keys())
                merged_sizes = list(type_quantity.values())
                total_quantity = sum(merged_sizes)

                percentages = [size / total_quantity * 100 for size in merged_sizes]

                colors = ['red', 'yellow', 'purple', 'orange', 'green', 'blue', 'cyan', 'magenta', 'lime', 'pink']
                fig, ax = Figure(figsize=(6, 2)), None
                ax = fig.add_subplot(111, aspect='equal')

                wedges, texts = ax.pie(merged_sizes, labels=None, startangle=90, colors=colors, pctdistance=0,)


                center_circle = plt.Circle((0, 0), 0.70, fc='white')
                ax.add_artist(center_circle)

                ax.set_title("Order Type")

                legend_labels = [f'{label} - {percentage:.2f}%' for label, percentage in zip(merged_labels, percentages)]
                ax.legend(wedges, legend_labels, title="Legend", bbox_to_anchor=(1, 0.5), loc="center left", borderaxespad=0.0)

                canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
                canvas_widget.get_tk_widget().place(x=550, y=350, width=600, height=250)
        except FileNotFoundError:
            pass

        try:
            with open("product_list.pkl", "rb") as product_file:
                product_data = pickle.load(product_file)
            with open("Sale_list.pkl", "rb") as Sale_file:
                Sale_data = pickle.load(Sale_file)
        except FileNotFoundError:
            return
        try:
            product_revenue_data = defaultdict(float)
            sale_revenue_data = defaultdict(float)
            
            for sale in Sale_data:
                matching_products = [product for product in product_data if product["Name"] == sale["Name"]]
                if matching_products:
                    product = matching_products[0]
                    product_revenue_data[product["Name"]] += product["Price"] * sale["Quantity"]
                    sale_revenue_data[sale["Name"]] += sale["Price"] * sale["Quantity"]

            products = list(product_revenue_data.keys())
            product_revenues = list(product_revenue_data.values())
            sale_revenues = [sale_revenue_data.get(product, 0) for product in products]

            fig, ax = plt.subplots(figsize=(5, 2))
            width = 0.2
            ind = range(len(products))

            ax.bar(ind, product_revenues, width, label='Product Revenue')
            ax.bar([i + width for i in ind], sale_revenues, width, label='Sale Revenue')

            ax.set_title('Product and Sale Revenue Comparison')
            ax.set_xticks([i + width/2 for i in ind])
            ax.set_xticklabels(products, rotation=0, ha='right', fontsize=8)
            ax.legend()

            for i, (product, product_revenue, sale_revenue) in enumerate(zip(products, product_revenues, sale_revenues)):
                ax.text(i, product_revenue + sale_revenue + 5, f"{product_revenue + sale_revenue:.2f}", ha='center', fontsize=6)

            canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
            canvas_widget.get_tk_widget().place(x=500, y=70, width=500, height=250)
            canvas_widget.draw()
        except FileNotFoundError:
            pass

        if self.history_tree is None:
            self.history_tree = ttk.Treeview(self.current_frame, columns=("Timestamp", "Action"), show="headings")
            self.history_tree.heading("Timestamp", text="Timestamp")
            self.history_tree.heading("Action", text="Action")

            vsb = Scrollbar(self.current_frame, orient=VERTICAL, command=self.history_tree.yview)
            self.history_tree.configure(yscrollcommand=vsb.set)

            self.history_tree.place(x=0, y=350, width=500, height=250)
            vsb.place(x=500, y=350, height=250)
        self.update_history_widget()

    def update_history_widget(self):
        if self.history_tree:
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)

            for entry in self.history_list:
                timestamp, _, action = entry.partition(" - ")  
                self.history_tree.insert("", "end", values=(timestamp, action))
        





class ProductPage(BasePage):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.product_list = []
        self.show_product_page()
        
    def show_product_page(self):
        self.load_data()
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = Frame(self.master, background="White")
        self.current_frame.place(x=200, y=100, width=1150, height=700)

        label_product = Label(self.current_frame, text="Product Page", font=("Arial", 24), background="White")
        label_product.pack(pady=20)

        low_quantity_products = [product for product in self.product_list if product["Quantity"] < 30]

        if low_quantity_products:
            message = "Low quantity alert!\n\nThe following products have a quantity lower than 30:\n"
            for product in low_quantity_products:
                message += f"{product['Name']} - {product['Quantity']} units\n"

            messagebox.showwarning("Low Quantity Alert", message)

        new_product_label = Button(self.current_frame, text="New Product", font=("Arial", 18),
                                   command=self.show_add_product_dialog)
        new_product_label.place(x=150, y=100, width=150, height=50)
        delete_button = Button(self.current_frame, text="Delete Selected Product", font=("Arial", 16),
                               command=self.delete_selected_product)
        delete_button.place(x=350, y=100, height=50)

        product_list_label = Label(self.current_frame, text="Product List", font=("Arial", 18), background="White")
        product_list_label.place(x=500, y=150, height=50)

        self.product_listbox = ttk.Treeview(self.current_frame, columns=("Name", "Price", "Quantity", "Type"),
                                            show="headings")
        self.product_listbox.heading("Name", text="Name")
        self.product_listbox.heading("Price", text="Price")
        self.product_listbox.heading("Quantity", text="Quantity")
        self.product_listbox.heading("Type", text="Type")
        self.product_listbox.place(x=150, y=200, height=300)
        scrollbar = Scrollbar(self.current_frame, orient="vertical", command=self.product_listbox.yview)
        scrollbar.place(x=935, y=200, height=300)
        self.product_list = [product for product in self.product_list if product["Quantity"] > 0]
        with open("product_list.pkl", "wb") as product_file:
            pickle.dump(self.product_list, product_file)
        self.product_listbox.configure(yscrollcommand=scrollbar.set)

        for product in self.product_list:
            self.product_listbox.insert("", "end", values=(product["Name"], product["Price"],product["Quantity"], product["Type"]))
            
        
    def show_add_product_dialog(self):
        add_product_window = Toplevel(self.master)
        add_product_window.geometry("400x400")          
        label_name = Label(add_product_window, text="Product Name:", font=("Arial", 16))
        label_name.pack()

        entry_name = Entry(add_product_window, font=("Arial", 16))
        entry_name.pack()

        label_price = Label(add_product_window, text="Price:", font=("Arial", 16))
        label_price.pack()

        entry_price = Entry(add_product_window, font=("Arial", 16))
        entry_price.pack()

        label_quantity = Label(add_product_window, text="Quantity:", font=("Arial", 16))
        label_quantity.pack()
        
        entry_quantity = Entry(add_product_window, font=("Arial", 16))
        entry_quantity.pack()
        label_Type = Label(add_product_window, text="Type:", font=("Arial", 16))
        label_Type.pack()
        entry_Type = ttk.Combobox(add_product_window, state="readonly")
        entry_Type["values"] = ("Food", "Toy", "Electronic", "Clothes", "Other")
        entry_Type.pack()

        add_button = Button(add_product_window, text="Add Product", font=("Arial", 16), command=lambda: self.add_product(add_product_window,entry_name.get(), entry_price.get(), entry_quantity.get(),entry_Type.get()))
        add_button.pack()
        

    def delete_selected_product(self):
        selected_items = self.product_listbox.selection()
        for item in selected_items:
            index = self.product_listbox.index(item)
            self.product_list.pop(index)
            self.product_listbox.delete(item)
        with open("product_list.pkl", "wb") as product_file:
            pickle.dump(self.product_list, product_file)
        self.show_product_page()

        self.add_to_history("Delete Product", item_name=index)

    
    def add_product(self, add_product_window, name, price, quantity, product_type):
        if name and price and quantity:
            try:
                price = float(price)
                quantity = int(quantity)

                existing_product = next((product for product in self.product_list if product["Name"] == name), None)

                if existing_product:
                    if existing_product["Price"] == price:
                        existing_product["Quantity"] += quantity
                    else:
                        new_name = self.generate_unique_product_name(name)
                        new_product_entry = {"Name": new_name, "Price": price, "Quantity": quantity, "Type": product_type}
                        self.product_listbox.insert("", "end", values=(new_name, price, quantity, product_type))
                        self.product_list.append(new_product_entry)
                else:
                    product_entry = {"Name": name, "Price": price, "Quantity": quantity, "Type": product_type}
                    self.product_listbox.insert("", "end", values=(name, price, quantity, product_type))
                    self.product_list.append(product_entry)

                self.product_listbox.update_idletasks()
                with open("product_list.pkl", "wb") as product_file:
                    pickle.dump(self.product_list, product_file)
                self.show_product_page()

                self.add_to_history("Added Product", item_name=name)

                add_product_window.destroy()
            except ValueError:
                pass

    def generate_unique_product_name(self, original_name):
        new_name = original_name
        count = 0
        while any(product["Name"] == new_name for product in self.product_list):
            count += 1
            new_name = f"{original_name} ({count})"
        return new_name



class OrderPage(BasePage):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.order_list = []  
        self.Sale_list = []
        self.order_list_modified = False
        self.show_order_page()

    def show_order_page(self):
        self.load_data()
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = Frame(self.master, background="White")
        self.current_frame.place(x=200, y=100, width=1150, height=700)

        label_order = Label(self.current_frame, text="Order Page", font=("Arial", 24), background="White")
        label_order.pack(pady=20)

        low_quantity_products = [product for product in self.product_list if product["Quantity"] < 30]

        if low_quantity_products:
            message = "Low quantity alert!\n\nThe following products have a quantity lower than 30:\n"
            for product in low_quantity_products:
                message += f"{product['Name']} - {product['Quantity']} units\n"

            messagebox.showwarning("Low Quantity Alert", message)

        new_order_label = Button(self.current_frame, text="New order", font=("Arial", 18), command=self.show_add_order_dialog)
        new_order_label.place(x=70, y=100, height=50)
        delete_button = Button(self.current_frame, text="Delete Selected Order", font=("Arial", 16),
                               command=self.delete_selected_order)
        delete_button.place(x=225, y=100, height=50)

        move_to_Sale_button = Button(self.current_frame, text="Order Complete", font=("Arial", 16),
                                     command=self.move_to_Sale)
        move_to_Sale_button.place(x=900, y=100, height=50)
        order_list_label = Label(self.current_frame, text="Order List", font=("Arial", 18), background="White")
        order_list_label.place(x=500, y=150, height=50)

        self.order_listbox = ttk.Treeview(self.current_frame, columns=("ID", "Name", "Price", "Quantity", "Payment"),
                                          show="headings")
        self.order_listbox.heading("ID", text="ID")
        self.order_listbox.heading("Name", text="Name")
        self.order_listbox.heading("Price", text="Price")
        self.order_listbox.heading("Quantity", text="Quantity")
        self.order_listbox.heading("Payment", text="Payment")

        self.order_listbox.place(x=70, y=200, height=300)
        scrollbar = Scrollbar(self.current_frame, orient="vertical", command=self.order_listbox.yview)
        scrollbar.place(x=1070, y=200, height=300)

        self.rearrange_order_ids()
        self.order_listbox.configure(yscrollcommand=scrollbar.set)
        for order in self.order_list:
            self.order_listbox.insert("", "end", values=(order["ID"], order["Name"], order["Price"], order["Quantity"], order["Payment"]))

    def rearrange_order_ids(self):
        for index, order in enumerate(self.order_list):
            order["ID"] = index + 1
        
    def show_add_order_dialog(self):
        add_order_window = Toplevel(self.current_frame)  
        add_order_window.geometry("400x400")
        next_id = 1
        if self.order_list:
            next_id = max(order["ID"] for order in self.order_list) + 1

        label_ID = Label(add_order_window, text="Order ID:", font=("Arial", 16))
        label_ID.pack()

        entry_ID = Entry(add_order_window, font=("Arial", 16))
        entry_ID.insert(0, next_id)  
        entry_ID.config(state='readonly') 
        entry_ID.pack()

        label_name = Label(add_order_window, text="Order Name:", font=("Arial", 16))
        label_name.pack()

        product_names = [product["Name"] for product in self.product_list]
        entry_name = ttk.Combobox(add_order_window, values=product_names, state="readonly", font=("Arial", 16))
        entry_name.pack()

        label_price = Label(add_order_window, text="Price:", font=("Arial", 16))
        label_price.pack()

        entry_price = Entry(add_order_window, font=("Arial", 16))
        entry_price.pack()

        label_quantity = Label(add_order_window, text="Quantity:", font=("Arial", 16))
        label_quantity.pack()

        entry_quantity = Entry(add_order_window, font=("Arial", 16))
        entry_quantity.pack()

        label_payment = Label(add_order_window, text="Payment:", font=("Arial", 16))
        label_payment.pack()

        entry_payment = ttk.Combobox(add_order_window, state="readonly")
        entry_payment["values"] = ("Cash", "Credit Card", "Debit Card")
        entry_payment.pack()


        add_button = Button(add_order_window, text="Add Order", font=("Arial", 16), command=lambda: self.add_order(add_order_window, entry_ID.get(), entry_name.get(), entry_price.get(), entry_quantity.get(), entry_payment.get()))
        add_button.pack()
        

    def add_order(self, add_order_window, id, name, price, quantity, payment):
        if name and price and quantity and payment:
            try:
                price = float(price)
                quantity = int(quantity)
                payment = str(payment)
            except ValueError:
                messagebox.showerror("Error", "Invalid input.")
                return
            next_id = 1
            if self.order_list:
                next_id = max(order["ID"] for order in self.order_list) + 1

            product_found = False
            for product in self.product_list:
                if product["Name"] == name and product["Quantity"] >= quantity:
                    product_found = True

                    if price < product["Price"]:
                        response = messagebox.askyesno("Warning", "Order price is lower than product price. Continue?")
                        if not response:
                            return 
                    break

            if product_found:
                for product in self.product_list:
                    if product["Name"] == name:
                        product["Quantity"] -= quantity  
                        break

                self.order_list.append({"ID": next_id, "Name": name, "Price": price, "Quantity": quantity, "Payment": payment})
                self.order_listbox.insert("", "end", values=(next_id, name, price, quantity, payment))
                self.order_listbox.update_idletasks()

                with open("product_list.pkl", "wb") as product_file:
                    pickle.dump(self.product_list, product_file)

                with open("order_list.pkl", "wb") as order_file:
                    pickle.dump(self.order_list, order_file)
                self.add_to_history("Added Order", item_name=name)
                with open("history_list.pkl", "wb") as history_file:
                    pickle.dump(self.history_list, history_file)
            else:
                messagebox.showerror("Error", "Product not in stock or insufficient quantity.")
        add_order_window.destroy()
        self.master.update()

    def delete_selected_order(self):
        selected_items = self.order_listbox.selection()
        for item in selected_items:
            index = self.order_listbox.index(item)
            order_item = self.order_list[index]
            id = order_item["ID"]
            name = order_item["Name"]
            price = order_item["Price"]
            quantity = order_item["Quantity"]
            payment = order_item["Payment"]
            self.return_product_to_inventory(id, name, price, quantity, payment)
            self.order_list.pop(index)
            self.order_listbox.delete(item)
        with open("order_list.pkl", "wb") as order_file:
            pickle.dump(self.order_list, order_file)
        self.add_to_history("Delete Order", item_name=index)
        with open("history_list.pkl", "wb") as history_file:
            pickle.dump(self.history_list, history_file)
        self.show_order_page()

    def return_product_to_inventory(self, id, name, price, quantity, payment):
        for product in self.product_list:
            if product["Name"] == name:
                product["Quantity"] += quantity
                
                break
        with open("product_list.pkl", "wb") as product_file:
            pickle.dump(self.product_list, product_file)
        messagebox.showinfo("Return to Product List", f"{quantity} units returned to product list.")


    def move_to_Sale(self):
        selected_items = self.order_listbox.selection()

        for item in selected_items:
            index = self.order_listbox.index(item)
            order_item = self.order_list[index]
            order_item["SaleDate"] = datetime.now().strftime("%Y-%m-%d")
            self.Sale_list.append(order_item)
            self.order_list.pop(index)
            self.order_listbox.delete(item)

        with open("Sale_list.pkl", "wb") as Sale_file:
            pickle.dump(self.Sale_list, Sale_file)
        
        with open("order_list.pkl", "wb") as order_file:
            pickle.dump(self.order_list, order_file)
        self.show_order_page()
        self.add_to_history("Order Complete")
        with open("history_list.pkl", "wb") as history_file:
            pickle.dump(self.history_list, history_file)


class SalePage(BasePage):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.Sale_list = []  
        self.show_Sale_page()

    def show_Sale_page(self):
        self.load_data()
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = Frame(self.master,background="White")
        self.current_frame.place(x=200, y=100, width=1150, height=600)

        label_Sale = Label(self.current_frame, text="Sale Page", font=("Arial", 24),background="White")
        label_Sale.pack(pady=20)

        self.Sale_listbox = ttk.Treeview(self.current_frame, columns=("Name", "Price", "Quantity", "Payment", "SaleDate"), show="headings")
        self.Sale_listbox.heading("Name", text="Name")
        self.Sale_listbox.heading("Price", text="Price")
        self.Sale_listbox.heading("Quantity", text="Quantity")
        self.Sale_listbox.heading("Payment", text="Payment")
        self.Sale_listbox.heading("SaleDate", text="SaleDate")

        self.Sale_listbox.column("Name", width=100)
        self.Sale_listbox.column("Price", width=100)
        self.Sale_listbox.column("Quantity", width=100)
        self.Sale_listbox.column("Payment", width=100)
        self.Sale_listbox.column("SaleDate", width=100)

        self.Sale_listbox.place(x=0, y=100, height=200)
        scrollbar = Scrollbar(self.current_frame, orient="vertical", command=self.Sale_listbox.yview)
        self.Sale_listbox.configure(yscrollcommand=scrollbar.set)
        scrollbar.place(x=500, y=100, height=200)
        for Sale in self.Sale_list:
            self.Sale_listbox.insert("", "end", values=(Sale["Name"], Sale["Price"], Sale["Quantity"], Sale["Payment"], Sale["SaleDate"]))
        
        self.generate_revenue_graph()
        self.generate_payment_method_donut()

    def calculate_revenue_over_time(self):
        revenue_data = defaultdict(float)

        with open("product_list.pkl", "rb") as product_file:
            product_data = pickle.load(product_file)

        for sale in self.Sale_list:
            sale_date = datetime.strptime(sale["SaleDate"], "%Y-%m-%d")
            sale_revenue = sale["Price"] * sale["Quantity"]

            product_revenue = sum(product["Price"] * sale["Quantity"] for product in product_data if product.get("Name") == sale["Name"])
            revenue_data[sale_date] += (sale_revenue - product_revenue)

        return revenue_data

    def generate_revenue_graph(self):
        try:
            revenue_data = self.calculate_revenue_over_time()

            if not revenue_data:
                return

            sorted_data = sorted(revenue_data.items(), key=lambda x: x[0])
            dates, revenues = zip(*sorted_data)

            last_7_days = 7
            if len(dates) > last_7_days:
                dates = dates[-last_7_days:]
                revenues = revenues[-last_7_days:]

            fig, ax = plt.subplots(figsize=(10, 10))
            ax.plot(dates, revenues, marker='o', linestyle='-')
            ax.set_title('Net Revenue Over the Last 7 Days')

            ax.xaxis.set_major_formatter(matplotlib.dates.DateFormatter('%Y-%m-%d'))
            plt.xticks(rotation=45, ha='right', fontsize=8)
            plt.yticks(fontsize=8)

            ax.set_xticks(dates)
            ax.set_xticklabels([date.strftime('%m-%d') for date in dates])

            canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
            canvas_widget.get_tk_widget().place(x=600, y=70, width=500, height=350)
            canvas_widget.draw()
        except FileNotFoundError:
            pass
    
    def generate_payment_method_donut(self):
        try:
            payment_data = defaultdict(float)

            for sale in self.Sale_list:
                payment_data[sale["Payment"]] += sale["Price"] * sale["Quantity"]

            if not payment_data:
                return

            payment_labels = list(payment_data.keys())
            payment_sizes = list(payment_data.values())
            total_revenue = sum(payment_sizes)

            percentages = [size / total_revenue * 100 for size in payment_sizes]

            colors = ['red', 'yellow', 'purple', 'orange', 'green', 'blue', 'cyan', 'magenta', 'lime', 'pink']
            fig, ax = plt.subplots(figsize=(6, 2))
            ax.pie(payment_sizes, labels=None, startangle=90, colors=colors, pctdistance=0.85)
            center_circle = plt.Circle((0, 0), 0.70, fc='white')
            ax.add_artist(center_circle)
            ax.set_title('Payment Methods')

            legend_labels = [f'{label} - {percentage:.2f}%' for label, percentage in zip(payment_labels, percentages)]
            ax.legend(legend_labels, title="Legend", bbox_to_anchor=(1, 0.5), loc="center left", borderaxespad=0.0)

            canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
            canvas_widget.get_tk_widget().place(x=-100, y=350, width=600, height=250)
            canvas_widget.draw()
        except FileNotFoundError:
            pass

class show_Report_page(BasePage):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.show_Report_page()

    def show_Report_page(self):
        if self.current_frame:
            self.current_frame.destroy()
        self.current_frame = Frame(self.master, background="White")
        self.current_frame.place(x=200, y=100, width=1150, height=600)
        label_Report = Label(self.current_frame, text="Report Page", font=("Arial", 24), background="White")
        label_Report.pack(pady=20)
        export_button = Button(self.current_frame, text="Export Sale Data", font=("Arial", 16), command=self.export_Sale_to_excel)
        export_button.place(x=50, y=100, width=200, height=50)
        Product_export_button = Button(self.current_frame, text="Export Product Data", font=("Arial", 16), command=self.export_product_to_excel)
        Product_export_button.place(x=50, y=200, width=200, height=50)

        try:
            with open("history_list.pkl", "rb") as history_file:
                self.history_list = pickle.load(history_file)
        except FileNotFoundError:
            return
        
        if self.history_tree is None:
            self.history_tree = ttk.Treeview(self.current_frame, columns=("Timestamp", "Action"), show="headings")
            self.history_tree.heading("Timestamp", text="Timestamp")
            self.history_tree.heading("Action", text="Action")

            vsb = Scrollbar(self.current_frame, orient=VERTICAL, command=self.history_tree.yview)
            self.history_tree.configure(yscrollcommand=vsb.set)

            self.history_tree.place(x=10, y=300, width=500, height=250)
            vsb.place(x=510, y=300, height=250)
        self.update_history_widget()

        try:
            with open("product_list.pkl", "rb") as product_file:
                product_data = pickle.load(product_file)
            with open("Sale_list.pkl", "rb") as Sale_file:
                Sale_data = pickle.load(Sale_file)
        except FileNotFoundError:
            return
        try:
            product_revenue_data = defaultdict(float)
            sale_revenue_data = defaultdict(float)
            
            for sale in Sale_data:
                matching_products = [product for product in product_data if product["Name"] == sale["Name"]]
                if matching_products:
                    product = matching_products[0]
                    product_revenue_data[product["Name"]] += product["Price"] * sale["Quantity"]
                    sale_revenue_data[sale["Name"]] += sale["Price"] * sale["Quantity"]

            products = list(product_revenue_data.keys())
            product_revenues = list(product_revenue_data.values())
            sale_revenues = [sale_revenue_data.get(product, 0) for product in products]

            fig, ax = plt.subplots(figsize=(5, 2))
            width = 0.2
            ind = range(len(products))

            ax.bar(ind, product_revenues, width, label='Product Revenue')
            ax.bar([i + width for i in ind], sale_revenues, width, label='Sale Revenue')

            ax.set_title('Product and Sale Revenue Comparison')
            ax.set_xticks([i + width/2 for i in ind])
            ax.set_xticklabels(products, rotation=0, ha='right', fontsize=8)
            ax.legend()

            for i, (product, product_revenue, sale_revenue) in enumerate(zip(products, product_revenues, sale_revenues)):
                ax.text(i, product_revenue + sale_revenue + 5, f"{product_revenue + sale_revenue:.2f}", ha='center', fontsize=6)

            canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
            canvas_widget.get_tk_widget().place(x=600, y=70, width=500, height=250)
            canvas_widget.draw()
        except FileNotFoundError:
            pass
        try:
            with open("Sale_list.pkl", "rb") as Sale_file:
                self.Sale_list = pickle.load(Sale_file)
            payment_data = defaultdict(float)

            for sale in self.Sale_list:
                payment_data[sale["Payment"]] += sale["Price"] * sale["Quantity"]

            if not payment_data:
                return

            payment_labels = list(payment_data.keys())
            payment_sizes = list(payment_data.values())
            total_revenue = sum(payment_sizes)

            percentages = [size / total_revenue * 100 for size in payment_sizes]

            colors = ['red', 'yellow', 'purple', 'orange', 'green', 'blue', 'cyan', 'magenta', 'lime', 'pink']
            fig, ax = plt.subplots(figsize=(6, 2))
            ax.pie(payment_sizes, labels=None, startangle=90, colors=colors, pctdistance=0.85)
            center_circle = plt.Circle((0, 0), 0.70, fc='white')
            ax.add_artist(center_circle)
            ax.set_title('Payment Methods')

            legend_labels = [f'{label} - {percentage:.2f}%' for label, percentage in zip(payment_labels, percentages)]
            ax.legend(legend_labels, title="Legend", bbox_to_anchor=(1, 0.5), loc="center left", borderaxespad=0.0)

            canvas_widget = FigureCanvasTkAgg(fig, master=self.current_frame)
            canvas_widget.get_tk_widget().place(x=540, y=350, width=600, height=250)
            canvas_widget.draw()

        except FileNotFoundError:
            pass
    def update_history_widget(self):
        if self.history_tree:
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)

            for entry in self.history_list:
                timestamp, _, action = entry.partition(" - ") 
                self.history_tree.insert("", "end", values=(timestamp, action))
    def export_Sale_to_excel(self):
        try:
            with open("Sale_list.pkl", "rb") as Sale_file:
                self.Sale_list = pickle.load(Sale_file)
            with open("product_list.pkl", "rb") as product_file:
                self.product_list = pickle.load(product_file)

            if not self.Sale_list:
                messagebox.showwarning("No Data", "No data available for export.")
                return

            for idx, sale in enumerate(self.Sale_list, start=1):
                sale["ID"] = idx

            df_sales = pd.DataFrame(self.Sale_list)

            total_revenue_dict = {product["Name"]: 0 for product in self.product_list}

            for sale in self.Sale_list:
                product_name = sale["Name"]
                revenue = sale["Price"] * sale["Quantity"]
                total_revenue_dict[product_name] += revenue
                sale["Revenue"] = revenue

            df_sales = pd.DataFrame(self.Sale_list)

            workbook = Workbook()
            worksheet = workbook.active

            header_sales = list(df_sales.columns)
            for col_num, header_text in enumerate(header_sales, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_text
                
            for row_num, row_data in enumerate(df_sales.values, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_data

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            if file_path:
                workbook.save(file_path)
                messagebox.showinfo("Export Successful", f"Data has been exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred during export:\n{str(e)}")
    def export_product_to_excel(self):
        try:
            with open("product_list.pkl", "rb") as product_file:
                self.product_list = pickle.load(product_file)

            if not self.product_list:
                messagebox.showwarning("No Data", "No data available for export.")
                return
            
            df_products = pd.DataFrame(self.product_list)

            workbook = Workbook()
            worksheet = workbook.active

            header_products = list(df_products.columns)
            for col_num, header_text in enumerate(header_products, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header_text

            for row_num, row_data in enumerate(df_products.values, 2):
                for col_num, cell_data in enumerate(row_data, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_data

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            if file_path:
                workbook.save(file_path)

                messagebox.showinfo("Export Successful", f"Data has been exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred during export:\n{str(e)}")
            

class InventoryManagementApp:
    def __init__(self):
        self.tk = Tk()
        self.tk.title("Inventory Management")
        self.tk.geometry("1350x700")
        self.tk.resizable(0, 0)

        Tako = PhotoImage(file="Tako.png")
        scaling_factor_x = 2
        scaling_factor_y = 2
        Tako = Tako.subsample(scaling_factor_x, scaling_factor_y)
        self.logo = Tako 
        self.create_menu()

        self.page_switcher = BasePage(self.tk)

        self.show_home_page()

    def create_menu(self):
        title = Label(self.tk, text="Inventory Management", font=("Arial", 30), background="cyan", padx=100, pady=10)
        title.place(x=0, y=0, width=1350, height=100)
        self.Menu = Label(self.tk , text="Menu", font=("Arial", 20))
        self.Menu.place(x=0, y=300, width=200, height=50) 
        logo_frame = LabelFrame(self.tk)
        logo_frame.place(x=0, y=100, width=200, height=200)
        logo_label = Label(logo_frame, image=self.logo, width=self.logo.width(), height=self.logo.height(), background="cyan")
        logo_label.pack()
        
        Home = Button(self.tk, text="Home", font=("Arial", 20), command=self.show_home_page, padx=100, pady=10)
        Home.place(x=0, y=350, width=200, height=50)
        Product = Button(self.tk, text="Product", font=("Arial", 20), command=self.show_product_page, padx=100, pady=10)
        Product.place(x=0, y=400, width=200, height=50)
        Order = Button(self.tk, text="Order", font=("Arial", 20), command=self.show_order_page, padx=100, pady=10)
        Order.place(x=0, y=450, width=200, height=50)
        Sale = Button(self.tk, text="Sale", font=("Arial", 20), command=self.show_Sale_page, padx=100, pady=10)
        Sale.place(x=0, y=500, width=200, height=50)
        Report = Button(self.tk, text="Report", font=("Arial", 20), command=self.show_Report_page, padx=100, pady=10)
        Report.place(x=0, y=550, width=200, height=50)
        Exit = Button(self.tk, text="Exit", font=("Arial", 20), command=self.on_closing, padx=100, pady=10)
        Exit.place(x=0, y=600, width=200, height=50)

    def show_home_page(self):
        self.page_switcher.switch_page(HomePage)

    def show_product_page(self):
        self.page_switcher.switch_page(ProductPage)
    
    def show_order_page(self):
        self.page_switcher.switch_page(OrderPage)

    def show_Sale_page(self):
        self.page_switcher.switch_page(SalePage)

    def show_Report_page(self):
        self.page_switcher.switch_page(show_Report_page)
    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.tk.destroy()
            self.tk.quit()
    def run(self):
        self.tk.mainloop()

if __name__ == "__main__":
    app = InventoryManagementApp()
    app.run()